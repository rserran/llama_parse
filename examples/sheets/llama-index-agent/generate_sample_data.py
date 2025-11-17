"""
Generate sample spreadsheets for LlamaSheets + LlamaIndex Agent workflows.

This script creates example Excel files that demonstrate different use cases:
1. Simple data table (for Workflow 1)
2. Regional sales data (for Workflow 2)
3. Complex budget with formatting (for Workflow 3)
4. Weekly sales report (for Workflow 4)

Usage:
    python generate_sample_data.py
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def generate_workflow_1_data(output_dir: Path) -> None:
    """Generate simple financial report for Workflow 1."""
    print("📊 Generating Workflow 1: financial_report_q1.xlsx")

    # Create sample quarterly data
    months = ["January", "February", "March"]
    categories = ["Revenue", "Cost of Goods Sold", "Operating Expenses", "Net Income"]

    data = []
    for category in categories:
        row: dict[str, str | int] = {"Category": category}
        for month in months:
            if category == "Revenue":
                value = random.randint(80000, 120000)
            elif category == "Cost of Goods Sold":
                value = random.randint(30000, 50000)
            elif category == "Operating Expenses":
                value = random.randint(20000, 35000)
            else:  # Net Income
                value = int(
                    int(row.get("January", 0))
                    + int(row.get("February", 0))
                    + int(row.get("March", 0))
                )
                value = random.randint(15000, 40000)
            row[month] = value
        data.append(row)

    df = pd.DataFrame(data)

    # Write to Excel
    output_file = output_dir / "financial_report_q1.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Q1 Summary", index=False)

        # Format it nicely
        worksheet = writer.sheets["Q1 Summary"]
        for cell in worksheet[1]:  # Header row
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)

    print(f"  ✅ Created {output_file}")


def generate_workflow_2_data(output_dir: Path) -> None:
    """Generate regional sales data for Workflow 2."""
    print("\n📊 Generating Workflow 2: Regional sales data")

    regions = ["northeast", "southeast", "west"]
    products = ["Widget A", "Widget B", "Widget C", "Gadget X", "Gadget Y"]

    for region in regions:
        data = []
        start_date = datetime(2024, 1, 1)

        # Generate 90 days of sales data
        for day in range(90):
            date = start_date + timedelta(days=day)
            # Random number of sales per day (3-8)
            for _ in range(random.randint(3, 8)):
                product = random.choice(products)
                units_sold = random.randint(1, 20)
                price_per_unit = random.randint(50, 200)
                revenue = units_sold * price_per_unit

                data.append(
                    {
                        "Date": date.strftime("%Y-%m-%d"),
                        "Product": product,
                        "Units_Sold": units_sold,
                        "Revenue": revenue,
                    }
                )

        df = pd.DataFrame(data)

        # Write to Excel
        output_file = output_dir / f"sales_{region}.xlsx"
        df.to_excel(output_file, sheet_name="Sales", index=False)
        print(f"  ✅ Created {output_file} ({len(df)} rows)")


def generate_workflow_3_data(output_dir: Path) -> None:
    """Generate complex budget spreadsheet with formatting for Workflow 3."""
    print("\n📊 Generating Workflow 3: company_budget_2024.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"

    # Define departments with colors
    departments = {
        "Engineering": "C6E0B4",
        "Marketing": "FFD966",
        "Sales": "F4B084",
        "Operations": "B4C7E7",
    }

    # Define categories
    categories = {
        "Personnel": ["Salaries", "Benefits", "Training"],
        "Infrastructure": ["Office Rent", "Equipment", "Software Licenses"],
        "Operations": ["Travel", "Supplies", "Miscellaneous"],
    }

    # Styles
    header_font = Font(bold=True, size=12)
    category_font = Font(bold=True, size=11)

    row = 1

    # Title
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "2024 Annual Budget"
    ws[f"A{row}"].font = Font(bold=True, size=14)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 2

    # Headers
    ws[f"A{row}"] = "Category"
    ws[f"B{row}"] = "Item"
    for i, dept in enumerate(departments.keys()):
        ws.cell(row, 3 + i, dept)
        ws.cell(row, 3 + i).font = header_font

    for cell in ws[row]:
        cell.font = header_font
    row += 1

    # Data
    for category, items in categories.items():
        # Category header (bold)
        ws[f"A{row}"] = category
        ws[f"A{row}"].font = category_font
        row += 1

        # Items with department budgets
        for item in items:
            ws[f"A{row}"] = ""
            ws[f"B{row}"] = item

            # Add budget amounts for each department (with color)
            for i, (dept, color) in enumerate(departments.items()):
                amount = random.randint(5000, 50000)
                cell = ws.cell(row, 3 + i, amount)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                cell.number_format = "$#,##0"

            row += 1

        row += 1  # Blank row between categories

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    for i in range(len(departments)):
        ws.column_dimensions[chr(67 + i)].width = 15  # C, D, E, F

    output_file = output_dir / "company_budget_2024.xlsx"
    wb.save(output_file)
    print(f"  ✅ Created {output_file}")
    print("     • Bold categories, colored departments, merged title cell")


def generate_workflow_4_data(output_dir: Path) -> None:
    """Generate weekly sales report for Workflow 4."""
    print("\n📊 Generating Workflow 4: sales_weekly.xlsx")

    products = [
        "Product A",
        "Product B",
        "Product C",
        "Product D",
        "Product E",
        "Product F",
        "Product G",
        "Product H",
    ]

    # Generate one week of data
    data = []
    start_date = datetime(2024, 11, 4)  # Monday

    for day in range(7):
        date = start_date + timedelta(days=day)
        # Each product has 3-10 transactions per day
        for product in products:
            for _ in range(random.randint(3, 10)):
                units = random.randint(1, 15)
                price = random.randint(20, 150)
                revenue = units * price

                data.append(
                    {
                        "Date": date.strftime("%Y-%m-%d"),
                        "Product": product,
                        "Units": units,
                        "Revenue": revenue,
                    }
                )

    df = pd.DataFrame(data)

    # Write to Excel with some formatting
    output_file = output_dir / "sales_weekly.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Weekly Sales", index=False)

        # Format header
        worksheet = writer.sheets["Weekly Sales"]
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    print(f"  ✅ Created {output_file} ({len(df)} rows)")


def main() -> None:
    """Generate all sample data files."""
    print("=" * 60)
    print("Generating Sample Data for LlamaSheets + Coding Agent Workflows")
    print("=" * 60)

    # Create output directory
    output_dir = Path("input_data")
    output_dir.mkdir(exist_ok=True)

    # Generate data for each workflow
    generate_workflow_1_data(output_dir)
    generate_workflow_2_data(output_dir)
    generate_workflow_3_data(output_dir)
    generate_workflow_4_data(output_dir)

    print("\n" + "=" * 60)
    print("✅ All sample data generated!")
    print("=" * 60)
    print(f"\nFiles created in {output_dir.absolute()}:")
    print("\nWorkflow 1 (Understanding a New Spreadsheet):")
    print("  • financial_report_q1.xlsx")
    print("\nWorkflow 2 (Generating Analysis Scripts):")
    print("  • sales_northeast.xlsx")
    print("  • sales_southeast.xlsx")
    print("  • sales_west.xlsx")
    print("\nWorkflow 3 (Using Cell Metadata):")
    print("  • company_budget_2024.xlsx")
    print("\nWorkflow 4 (Complete Automation):")
    print("  • sales_weekly.xlsx")
    print("\nYou can now use these files with the workflows in the documentation!")


if __name__ == "__main__":
    main()
